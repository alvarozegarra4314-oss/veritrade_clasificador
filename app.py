import sys
import warnings
import threading
from pathlib import Path
from io import BytesIO
from datetime import datetime
import pandas as pd
import streamlit as st

# ---------------------------------------------------------------------
# CONFIGURACIÓN DE PÁGINA Y ESTILOS
# ---------------------------------------------------------------------
st.set_page_config(page_title="Clasificador de Importaciones — Veritrade", page_icon="🗂️", layout="wide")

# Ignorar la advertencia de obsolescencia de la librería de Gemini
warnings.filterwarnings("ignore", category=FutureWarning, module="google.generativeai")

# CSS personalizado para emular el diseño web (Botón principal grande y métricas con fondo)
st.markdown("""
<style>
    /* Estilizar el botón principal de procesar */
    .stButton>button[kind="primary"],
    div[data-testid="stBaseButton-primary"] {
        height: 3.5rem;
        font-size: 1.2rem;
        font-weight: bold;
        border-radius: 0.5rem;
    }
    /* Tarjetas suaves para las métricas (testid vigente en Streamlit >= 1.29) */
    div[data-testid="stMetric"] {
        background-color: #f8fafc;
        border: 1px solid #e2e8f0;
        padding: 1rem;
        border-radius: 0.5rem;
        box-shadow: 0 1px 2px 0 rgba(0, 0, 0, 0.05);
    }
</style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------------------
# IMPORTACIONES LOCALES
# ---------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
SRC_DIR = BASE_DIR / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from src.pipeline import procesar_dataframe_dinamico
from src.maestro.loader import CargarMaestro
from src.excel_io import sanitizar_dataframe_para_excel
from src.ia_rescate import RescatadorIA, GENAI_DISPONIBLE
from src.maestro_optimizer import guardar_maestro_optimizado
from src.texto_utils import identificar_columnas_descripcion
from src.excel_estilos import aplicar_estilo_hoja_excel
from src import config

# ---------------------------------------------------------------------
# 0. INICIALIZACIÓN DE VARIABLES DE SESIÓN (SESSION STATE)
# ---------------------------------------------------------------------
if "df_resultado" not in st.session_state:
    st.session_state.df_resultado = None
if "df_export_data" not in st.session_state:
    st.session_state.df_export_data = None
if "maestro_opt_data" not in st.session_state:
    st.session_state.maestro_opt_data = None
if "resumen_opt" not in st.session_state:
    st.session_state.resumen_opt = None
if "linea_producto" not in st.session_state:
    st.session_state.linea_producto = "Producto"
if "proceso_completado" not in st.session_state:
    st.session_state.proceso_completado = False
if "kpis" not in st.session_state:
    st.session_state.kpis = {}
if "df_pendientes" not in st.session_state:
    st.session_state.df_pendientes = None
if "linea_detectada" not in st.session_state:
    st.session_state.linea_detectada = "Producto"
if "archivo_origen" not in st.session_state:
    st.session_state.archivo_origen = ""
if "hoja_origen" not in st.session_state:
    st.session_state.hoja_origen = ""
if "modelo_ia_usado" not in st.session_state:
    st.session_state.modelo_ia_usado = ""
if "_usar_ia" not in st.session_state:
    st.session_state._usar_ia = False
if "var_principal_nombre" not in st.session_state:
    st.session_state.var_principal_nombre = ""
if "valor_principal" not in st.session_state:
    st.session_state.valor_principal = ""
if "variables_categoricas" not in st.session_state:
    st.session_state.variables_categoricas = []
if "variables_potencia" not in st.session_state:
    st.session_state.variables_potencia = []

# Forzar uso del modelo actual de config (evita que sesiones viejas usen modelos obsoletos)
if "_modelo_ia_forzado" not in st.session_state:
    st.session_state._modelo_ia_forzado = config.MODELO_IA_DEFAULT
elif st.session_state._modelo_ia_forzado != config.MODELO_IA_DEFAULT:
    st.session_state._modelo_ia_forzado = config.MODELO_IA_DEFAULT
    st.session_state.modelo_ia_usado = ""
if "processing_active" not in st.session_state:
    st.session_state.processing_active = False
if "processing_done" not in st.session_state:
    st.session_state.processing_done = False
if "progress_pct" not in st.session_state:
    st.session_state.progress_pct = 0.0
if "progress_text" not in st.session_state:
    st.session_state.progress_text = ""
if "progress_error" not in st.session_state:
    st.session_state.progress_error = None

def _obtener_api_key_de_secrets() -> str:
    try:
        return st.secrets.get("GEMINI_API_KEY", "").strip()
    except Exception:
        return ""


@st.cache_data(show_spinner=False)
def _probar_api_key(api_key: str, modelo: str):
    """Llamada mínima a Gemini para validar la key ANTES de una corrida larga.
    Devuelve (ok: bool, detalle: str). Cacheada por (key, modelo) para no
    repetir la llamada si el usuario vuelve a pulsar con los mismos valores."""
    try:
        from google import genai
        api_key = api_key.strip()
        cliente = genai.Client(api_key=api_key)
        respuesta = cliente.models.generate_content(
            model=modelo,
            contents="Responde únicamente con la palabra OK.",
        )
        texto = (respuesta.text or "").strip()
        return True, f"✅ Conexión exitosa con **{modelo}**. El modelo respondió: '{texto[:20]}'"
    except Exception as e:
        detalle = str(e)
        if "429" in detalle or "quota" in detalle.lower():
            detalle += "\n\nLa key es válida pero no tiene cuota disponible en este momento."
        return False, f"❌ No se pudo conectar con **{modelo}**: {detalle}"

# Valores que el motor considera "marca sin resolver" (defaults del maestro)
VALORES_MARCA_SIN_RESOLVER = {
    "MARCA GENERICA", "MARCA PRINCIPAL", "MARCA COMPONENTES",
    "S/M", "SIN MARCA", "GENERICO", "NO APLICA",
}


@st.cache_data(show_spinner=False)
def _listar_hojas_y_filas(bytes_raw: bytes):
    """Lista las hojas de un Excel, su cantidad de filas y columnas con nombre (rápido)."""
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(bytes_raw), read_only=True, data_only=True)
    info = []
    for ws in wb.worksheets:
        # Leer solo la primera fila para contar columnas con nombre real
        cols_nombradas = 0
        for fila in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            cols_nombradas = sum(1 for v in fila if v is not None and not str(v).startswith("Unnamed"))
        info.append({"nombre": ws.title, "filas": int(ws.max_row or 0), "cols_nombradas": cols_nombradas})
    wb.close()
    return info


def _hoja_recomendada(info_hojas):
    """
    Preselecciona la hoja de datos real:
    prioriza nombres tipo 'Veritrade'/'data'/'2025'/'2024' CON columnas
    nombradas (descarta hojas auxiliares de gráficos que salen 'Unnamed').
    """
    if not info_hojas:
        return None

    def peso(item):
        nombre = item["nombre"].upper()
        es_datos = any(k in nombre for k in ("VERITRADE", "DATA", "2025", "2024"))
        tiene_cols = item.get("cols_nombradas", 0) >= 3
        bonus = 10_000_000 if (es_datos and tiene_cols) else (1_000_000 if es_datos else 0)
        return bonus + item["filas"]

    return max(info_hojas, key=peso)["nombre"]


def _cargar_maestro_incluido():
    """Devuelve (bytes, nombre) del maestro incluido en el proyecto, o (None, None)."""
    ruta = BASE_DIR / "data" / "maestro" / "Maestro_UPS_v2.xlsx"
    try:
        if ruta.exists():
            return ruta.read_bytes(), ruta.name
    except Exception:
        pass
    return None, None


def ejecutar_pipeline_reglas_cached(bytes_raw: bytes, bytes_maestro: bytes, hoja: str):
    """Lee el Excel crudo y el maestro SIN caché de Streamlit (no se puede
    usar @st.cache_data dentro de un hilo background). El overhead de
    re-leer el Excel es mínimo comparado con el pipeline completo."""
    df_raw = pd.read_excel(BytesIO(bytes_raw), sheet_name=hoja, engine="openpyxl")
    maestro = CargarMaestro(BytesIO(bytes_maestro))
    return df_raw, maestro


@st.cache_data(show_spinner=False)
def _vista_previa_cruda(bytes_raw: bytes, hoja: str) -> pd.DataFrame:
    """Encabezados + 5 filas de una hoja. Cacheado para no re-parsear el
    Excel completo en cada rerun (la validación de columnas corre aquí)."""
    return pd.read_excel(BytesIO(bytes_raw), sheet_name=hoja, nrows=5)


def _mensaje_columnas_no_reconocidas(columnas) -> str:
    """Mensaje accionable cuando la hoja no tiene columnas de descripción."""
    lista_cols = ", ".join(map(str, list(columnas)[:15])) + (" ..." if len(columnas) > 15 else "")
    return (
        "⚠️ La hoja seleccionada no tiene ninguna columna de descripción reconocible, "
        "así que no se puede clasificar.\n\n"
        "**Qué buscamos:** columnas cuyo nombre contenga *DESCRIPCION*, *DETALLE*, "
        "*MERCADERIA* o *COMMODITY*. Las administrativas (*PARTIDA*, *ARANCEL*, "
        "*NANDINA*, *SUBPARTIDA*) se ignoran a propósito.\n\n"
        f"**Columnas encontradas:** {lista_cols}\n\n"
        "Selecciona otra hoja aquí arriba o verifica que el archivo sea el export "
        "Veritrade correcto."
    )

# =====================================================================
# ENCABEZADO
# =====================================================================
st.title("🗂️ Clasificador de Importaciones — Veritrade")
st.markdown("Sube tu archivo de importaciones y obtén la clasificación por producto y marca. **No necesitas saber de reglas:** la herramienta aplica el maestro de la línea automáticamente y, si activas la IA, rescata lo que no logra resolver.")
st.write("") # Espaciador

# =====================================================================
# TABS PRINCIPALES
# =====================================================================
tab_clasificar, tab_crear = st.tabs(["📊 Clasificar Importaciones", "🔧 Crear Maestro"])

with tab_clasificar:
    # =====================================================================
    # SECCIÓN 1: CARGA DE ARCHIVOS
    # =====================================================================
    c_raw, c_maestro = st.columns(2)

    hoja_raw_valida = False
    archivo_raw = None
    maestro_bytes = None
    maestro_nombre = None
    maestro_info = None

    with c_raw:
        with st.container(border=True):
            st.subheader("1. Archivo de Datos Crudos")
            st.caption("Sube el archivo Excel con las descripciones a analizar.")
            archivo_raw = st.file_uploader("Arrastra tu archivo .xlsx aquí", type=["xlsx"], label_visibility="collapsed")

            if archivo_raw is not None:
                try:
                    archivo_raw.seek(0)
                    info_hojas = _listar_hojas_y_filas(archivo_raw.getvalue())
                    nombres_hojas = [i["nombre"] for i in info_hojas]

                    if nombres_hojas:
                        recomendada = _hoja_recomendada(info_hojas)
                        idx_default = nombres_hojas.index(recomendada) if recomendada in nombres_hojas else 0
                        hoja_raw = st.selectbox(
                            "Hoja a procesar",
                            nombres_hojas,
                            index=idx_default,
                            help="Se preseleccionó automáticamente la hoja con más datos.",
                        )
                        filas_estimadas = info_hojas[idx_default]["filas"]
                        cols_nombradas = info_hojas[idx_default].get("cols_nombradas", 0)
                        st.caption(f"📄 Hoja **{hoja_raw}** — ~{filas_estimadas:,} filas · {cols_nombradas} columnas")

                        # Validación temprana: la hoja debe tener columnas de
                        # descripción reconocibles ANTES de permitir procesar.
                        # Así el usuario descubre el problema aquí y no tras un error.
                        df_prev = None
                        try:
                            df_prev = _vista_previa_cruda(archivo_raw.getvalue(), hoja_raw)
                        except Exception as e:
                            st.error(f"No se pudo leer la hoja '{hoja_raw}': {e}")

                        if df_prev is not None:
                            cols_desc_detectadas = identificar_columnas_descripcion(df_prev.columns)
                            if cols_desc_detectadas:
                                hoja_raw_valida = True
                                st.caption(f"🔎 Columnas de descripción detectadas: **{', '.join(cols_desc_detectadas)}**")
                                with st.expander("👀 Vista previa del archivo crudo"):
                                    st.caption(f"Columnas detectadas: {len(df_prev.columns)}")
                                    st.dataframe(df_prev, width="stretch", hide_index=True)
                            else:
                                hoja_raw_valida = False
                                st.error(_mensaje_columnas_no_reconocidas(df_prev.columns))
                    else:
                        st.error("El archivo no contiene hojas.")
                        hoja_raw = None
                except Exception as e:
                    st.error(f"No se pudo leer el archivo: {e}")
                    hoja_raw = None
            else:
                hoja_raw = None

    with c_maestro:
        with st.container(border=True):
            st.subheader("2. Maestro de Reglas")
            st.caption("Usa siempre tu propio maestro .xlsx. La conexión local del proyecto ya no es necesaria.")

            archivo_maestro_up = st.file_uploader(
                "Arrastra tu archivo maestro .xlsx aquí",
                type=["xlsx"],
                label_visibility="collapsed",
                key="up_maestro",
            )
            if archivo_maestro_up is not None:
                archivo_maestro_up.seek(0)
                maestro_bytes = archivo_maestro_up.getvalue()
                maestro_nombre = archivo_maestro_up.name
            else:
                maestro_bytes = None
                maestro_nombre = None

            if maestro_bytes:
                try:
                    maestro_info = CargarMaestro(ruta_excel=BytesIO(maestro_bytes))
                    linea_detectada = maestro_info.config_linea.get("LINEA_PRODUCTO", "Producto")
                    n_cond = len(getattr(maestro_info, "condicionales", []))
                    st.session_state.linea_detectada = linea_detectada
                    # Guardamos la variable principal y su valor para poder
                    # mostrarlos con nombre real en los KPIs de resultados.
                    st.session_state.var_principal_nombre = maestro_info.variable_producto_principal
                    st.session_state.valor_principal = maestro_info.valor_producto_principal
                    # Guardamos las variables categóricas (características no numéricas)
                    # y las de potencia (numéricas) para poder calcular KPIs de
                    # "característica identificada" en los resultados.
                    st.session_state.variables_categoricas = list(getattr(maestro_info, "variables_categoricas", []))
                    st.session_state.variables_potencia = list(getattr(maestro_info, "variables_potencia", []))
                    st.success(f"✅ **Maestro:** {maestro_nombre} | **Línea:** {linea_detectada} | **Reglas activas:** {n_cond}")
                except Exception as e:
                    st.warning(f"Error al leer el maestro: {e}")
            else:
                st.info("📥 Sube tu maestro propio para habilitar el análisis.")

    # =====================================================================
    # SECCIÓN 2: CONFIGURACIÓN DE IA
    # =====================================================================
    st.write("") # Espaciador
    with st.container(border=True):
        st.markdown("### 🤖 Rescate por IA Generativa")
        st.caption("Delega a Gemini el análisis de las descripciones que el motor de reglas no logre resolver. *(Opcional pero recomendado)*.")
    
        if not GENAI_DISPONIBLE:
            st.error("⚠️ El paquete 'google-genai' no está instalado. Instálalo con: pip install google-genai")
            usar_ia = False
            api_key = ""
            modelo_ia = config.MODELO_IA_DEFAULT
            rpm_limite = 12
        else:
            usar_ia = st.toggle(
                f"Activar motor de rescate por IA (Gemini · {config.MODELO_IA_DEFAULT})",
                value=False,
            )

            api_key = None
            modelo_ia = config.MODELO_IA_DEFAULT
            rpm_limite = 12  # default cuando IA está desactivada
            if usar_ia:
                c_key, c_rpm = st.columns([2, 1])
                with c_key:
                    api_key = st.text_input(
                        "API Key",
                        type="password",
                        value=_obtener_api_key_de_secrets(),
                        help="Se toma de .streamlit/secrets.toml si existe; si no, pégala aquí.",
                    ).strip()
                    if not api_key:
                        st.warning("Se requiere API Key de Gemini.")
                with c_rpm:
                    rpm_limite = st.slider("Límite de Peticiones (RPM)", min_value=1, max_value=60, value=12)

                c_modelo, c_test = st.columns([2, 1])
                with c_modelo:
                    modelo_ia = st.selectbox(
                        "Modelo de IA",
                        config.MODELOS_IA_DISPONIBLES,
                        index=config.MODELOS_IA_DISPONIBLES.index(config.MODELO_IA_DEFAULT),
                        help="Si Google retira o renombra un modelo, elige otro de la lista sin cambiar código.",
                    )
                with c_test:
                    st.write("")  # alinear con el selectbox
                    probar_conexion = st.button("🔌 Probar conexión", width="stretch")

                if probar_conexion:
                    if api_key:
                        with st.spinner("Probando conexión con Gemini..."):
                            ok, detalle = _probar_api_key(api_key, modelo_ia)
                        if ok:
                            st.success(detalle)
                        else:
                            st.error(detalle)
                    else:
                        st.warning("Ingresa una API Key primero.")

def _generar_excel_resultado(df_resultado, kpis, linea, archivo_origen, hoja_origen, modelo_ia_usado):
    """Genera el buffer Excel (Resumen + Clasificación) de forma lazy.
    Solo se ejecuta cuando el usuario pulsa descargar, NO durante el procesamiento.
    Esto evita que el hilo se bloquee 10-30s generando openpyxl para 14k+ filas."""
    _columnas_auxiliares_no_exportar = {
        "Marca_Declarada", "Tipo_Producto_Detallado",
        "Producto_Texto_Desc1", "Modelo_Serie_Desc1",
        "Rescatado_Por_IA",
    }
    _df_export = df_resultado.drop(
        columns=[c for c in df_resultado.columns if c in _columnas_auxiliares_no_exportar],
        errors="ignore",
    )
    _df_export = sanitizar_dataframe_para_excel(_df_export)
    _total_filas = len(df_resultado)
    _cobertura_pct = f"{kpis.get('con_producto', 0) / max(_total_filas, 1):.1%}"
    _df_resumen = pd.DataFrame([
        ("Fecha de proceso", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Archivo origen", archivo_origen),
        ("Hoja procesada", hoja_origen),
        ("Línea de producto", linea),
        ("Motor de clasificación", modelo_ia_usado),
        ("Total de filas", f"{kpis.get('total', 0):,}"),
        ("Con producto identificado", f"{kpis.get('con_producto', 0):,} ({_cobertura_pct})"),
        ("Sin producto identificado", f"{kpis.get('sin_producto', 0):,}"),
        ("Sin marca (genérica)", f"{kpis.get('sin_marca', 0):,}"),
        ("Pendientes de revisión", f"{kpis.get('pendientes', 0):,}"),
        ("Rescatados por IA", f"{kpis.get('rescatados', 0):,}"),
        ("Resueltos desde caché IA (ahorro)", f"{kpis.get('cache', 0):,}"),
        ("Nuevas reglas aprendidas", f"+{kpis.get('nuevas', 0)}"),
        ("Errores de IA", f"{kpis.get('errores', 0):,}"),
    ], columns=["Parametro", "Valor"])

    output_buf = BytesIO()
    with pd.ExcelWriter(output_buf, engine="openpyxl") as writer:
        _df_resumen.to_excel(writer, index=False, sheet_name="Resumen")
        _df_export.to_excel(writer, index=False, sheet_name="Clasificacion")
        for nh, dh in (("Resumen", _df_resumen), ("Clasificacion", _df_export)):
            try:
                aplicar_estilo_hoja_excel(writer.sheets[nh], dh)
            except Exception:
                pass
    return output_buf.getvalue()


# =====================================================================
# SECCIÓN 3: ACCIÓN PRINCIPAL (PROCESAMIENTO)
# =====================================================================
st.write("")
listo_para_procesar = (
    archivo_raw is not None and maestro_bytes is not None and hoja_raw_valida
    and (not usar_ia or api_key)
    and not st.session_state.get("processing_active", False)
)

procesar = st.button(
    "▶️ PROCESAR CLASIFICACIÓN",
    type="primary",
    width="stretch",
    disabled=not listo_para_procesar,
)

if procesar:
    if st.session_state.get("processing_active"):
        st.warning("⏳ Ya hay un procesamiento en curso. Espera a que termine.")
    else:
        linea = st.session_state.get("linea_detectada", "Producto")
        archivo_raw.seek(0)
        df_raw_bytes = archivo_raw.getvalue()

        st.session_state.df_raw_bytes = df_raw_bytes
        st.session_state.maestro_bytes = maestro_bytes
        st.session_state.hoja_raw = hoja_raw
        st.session_state.linea_producto = linea
        st.session_state.archivo_origen = archivo_raw.name
        st.session_state.hoja_origen = hoja_raw
        st.session_state._usar_ia = bool(usar_ia)
        st.session_state.modelo_ia_usado = (
            f"Reglas + IA ({modelo_ia})" if usar_ia else "Reglas deterministas (sin IA)"
        )

        _shared = {
            "progress_pct": 0.0,
            "progress_text": "Preparando procesamiento...",
            "progress_error": None,
            "done": False,
            "result": None,
        }
        st.session_state._thread_shared = _shared
        st.session_state.processing_active = True
        st.session_state.processing_done = False
        st.session_state.proceso_completado = False
        st.session_state._rerun_triggered = False

        def _procesar_en_hilo(_shared, _raw_bytes, _maestro_bytes, _hoja, _usar_ia, _api_key, _rpm, _modelo):
            """Wrapper que ejecuta el pipeline completo en un hilo background.
            Solo escribe en _shared (dict), NUNCA en st.session_state."""
            try:
                _df_raw, _maestro = ejecutar_pipeline_reglas_cached(
                    _raw_bytes, _maestro_bytes, _hoja
                )

                _rescatador = None
                if _usar_ia and _api_key:
                    _rescatador = RescatadorIA(
                        api_key=_api_key, maestro=_maestro, rpm_limite=_rpm, modelo=_modelo
                    )

                def _cb_progreso(fase, i, total):
                    if not total or total <= 0:
                        return
                    pct = min(i / total, 1.0)
                    if fase == "reglas":
                        txt = f"Fase 1/2 · Reglas: {i:,} de {total:,} filas"
                    else:
                        txt = f"Fase 2/2 · IA: {i:,} de {total:,} descripciones"
                    _shared["progress_pct"] = pct
                    _shared["progress_text"] = txt

                try:
                    _df_resultado = procesar_dataframe_dinamico(
                        _df_raw, _maestro, rescatador_ia=_rescatador, progreso_callback=_cb_progreso
                    )
                finally:
                    if _rescatador is not None:
                        _rescatador.cerrar()

                if _rescatador is not None:
                    _desde_cache = (
                        _rescatador.descripciones_desde_cache_mem
                        + _rescatador.descripciones_desde_cache_db
                    )
                    _via_api = _rescatador.descripciones_rescatadas_api
                    if _via_api == 0 and _desde_cache == 0:
                        _texto_final = "✅ Completado · Las reglas resolvieron todo"
                    elif _via_api == 0:
                        _texto_final = (
                            f"✅ Completado · Caché IA: {_desde_cache:,} (sin gastar cuota)"
                        )
                    else:
                        _texto_final = (
                            f"✅ Completado · IA rescató {_via_api:,} "
                            f"(+{_desde_cache:,} desde caché)"
                        )
                else:
                    _texto_final = "✅ Completado · Solo reglas deterministas"

                _shared["progress_pct"] = 1.0
                _shared["progress_text"] = _texto_final

                _total_filas = len(_df_resultado)
                _kpis = {
                    "total": _total_filas, "rescatados": 0, "cache": 0, "nuevas": 0, "errores": 0,
                    "con_producto": 0, "sin_producto": 0, "sin_marca": 0, "pendientes": 0,
                }

                if "Producto_Declarado" in _df_resultado:
                    _kpis["con_producto"] = int(_df_resultado["Producto_Declarado"].notna().sum())
                    _kpis["sin_producto"] = int(_df_resultado["Producto_Declarado"].isna().sum())

                if "Marca_Extraida" in _df_resultado:
                    _kpis["sin_marca"] = int(
                        _df_resultado["Marca_Extraida"].astype(str).str.upper().isin(VALORES_MARCA_SIN_RESOLVER).sum()
                    )

                _pend_mask = pd.Series(False, index=_df_resultado.index)
                if "Producto_Declarado" in _df_resultado:
                    _pend_mask |= _df_resultado["Producto_Declarado"].isna()
                if "Marca_Extraida" in _df_resultado:
                    _pend_mask |= _df_resultado["Marca_Extraida"].astype(str).str.upper().isin(VALORES_MARCA_SIN_RESOLVER)
                _kpis["pendientes"] = int(_pend_mask.sum())

                if _rescatador is not None:
                    _kpis["rescatados"] = _rescatador.descripciones_rescatadas_api
                    _kpis["cache"] = _rescatador.llamadas_desde_cache
                    _kpis["errores"] = _rescatador.errores

                _maestro_opt_data = None
                _resumen_opt = None
                if _rescatador is not None:
                    _propuestas = getattr(_rescatador, "propuestas_aprendizaje", None)
                    if _propuestas and (_propuestas.get("nuevas_marcas") or _propuestas.get("nuevas_caracteristicas")):
                        _buf_opt = BytesIO()
                        _resumen_opt = guardar_maestro_optimizado(
                            ruta_maestro_original=BytesIO(_maestro_bytes), propuestas=_propuestas, ruta_salida=_buf_opt
                        )
                        _maestro_opt_data = _buf_opt.getvalue()
                        _kpis["nuevas"] = _resumen_opt.get("marcas_agregadas", 0) + _resumen_opt.get("caracteristicas_agregadas", 0)

                _shared["df_resultado"] = _df_resultado
                _shared["df_pendientes"] = _df_resultado[_pend_mask].copy()
                _shared["kpis"] = _kpis
                _shared["maestro_opt_data"] = _maestro_opt_data
                _shared["resumen_opt"] = _resumen_opt

            except Exception as e:
                _shared["progress_error"] = str(e)
                _shared["progress_text"] = f"❌ Error: {e}"
            finally:
                _shared["done"] = True

        _thread = threading.Thread(
            target=_procesar_en_hilo,
            args=(_shared, df_raw_bytes, maestro_bytes, hoja_raw, usar_ia, api_key, rpm_limite, modelo_ia),
            daemon=True,
        )
        _thread.start()


# =====================================================================
# SECCIÓN 3b: INDICADOR DE PROCESAMIENTO (se muestra en reruns posteriores)
# =====================================================================
@st.fragment(run_every="500ms")
def _fragmento_progreso_rerun():
    _sh = st.session_state.get("_thread_shared")
    if _sh is None or not st.session_state.get("processing_active", False):
        return

    st.session_state.progress_pct = _sh.get("progress_pct", 0.0)
    st.session_state.progress_text = _sh.get("progress_text", "Iniciando...")
    st.session_state.progress_error = _sh.get("progress_error")

    if _sh.get("done"):
        st.session_state.df_resultado = _sh.get("df_resultado")
        st.session_state.df_pendientes = _sh.get("df_pendientes")
        st.session_state.kpis = _sh.get("kpis")
        st.session_state.maestro_opt_data = _sh.get("maestro_opt_data")
        st.session_state.resumen_opt = _sh.get("resumen_opt")
        st.session_state.df_export_data = None
        st.session_state.proceso_completado = True
        st.session_state.processing_active = False
        st.session_state.processing_done = True
        if not st.session_state.get("_rerun_triggered"):
            st.session_state._rerun_triggered = True
            st.rerun()
        return

    pct = st.session_state.get("progress_pct", 0.0)
    texto = st.session_state.get("progress_text", "Iniciando...")
    st.progress(pct, text=texto)


if st.session_state.get("processing_active", False):
    st.write("")
    _fragmento_progreso_rerun()

# =====================================================================
# SECCIÓN 4: ÁREA DE RESULTADOS (PERSISTENTE)
# =====================================================================
if st.session_state.get("proceso_completado") and st.session_state.df_resultado is not None:
    st.write("")
    st.divider()

    col_titulo, col_tag = st.columns([4, 1])
    with col_titulo:
        st.markdown("### 📥 Resultados y Descargas")
    with col_tag:
        st.success("✅ Proceso Finalizado")

    kpis = st.session_state.kpis
    total = max(kpis.get("total", 1), 1)
    usar_ia = st.session_state.get("_usar_ia", False)

    # ---- Bloque 1: Resultado del proceso ----
    if usar_ia:
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total Filas", f"{kpis.get('total', 0):,}")
        m2.metric("Rescatados IA", f"{kpis.get('rescatados', 0):,}")
        m3.metric("Ahorro Caché", f"{kpis.get('cache', 0):,}")
        m4.metric("Nuevas Reglas", f"+{kpis.get('nuevas', 0)}")
    else:
        m1, m2 = st.columns(2)
        m1.metric("Total Filas", f"{kpis.get('total', 0):,}")
        m2.metric("Motor", "Reglas deterministas")

    if kpis.get("errores", 0) > 0:
        st.warning(f"⚠️ {kpis['errores']} descripciones tuvieron errores de conexión con Gemini.")

    st.write("")

    # ---- Bloque 2: Cobertura de clasificación y marcas identificadas ----
    con_producto = kpis.get("con_producto", 0)
    pct_con_producto = con_producto / total

    # Barra de cobertura general: % de filas donde se identificó ALGÚN tipo de producto
    st.markdown("#### 🎯 Cobertura de clasificación")
    st.caption("Porcentaje de filas donde el motor logró identificar el tipo de producto (cualquiera: UPS, interruptor, batería, etc.).")
    st.progress(
        min(pct_con_producto, 1.0),
        text=f"Identificación de producto: {con_producto:,} de {total:,} filas ({pct_con_producto:.1%})",
    )

    # Marcas identificadas: número de marcas distintas reales (excluye genéricas/sin marca)
    df_res = st.session_state.df_resultado
    if "Marca_Extraida" in df_res.columns:
        marcas_unicas = df_res["Marca_Extraida"].dropna().astype(str).str.strip().str.upper()
        mask_marca_real = ~marcas_unicas.isin(VALORES_MARCA_SIN_RESOLVER)
        n_filas_marca_real = int(mask_marca_real.sum())
        n_marcas_identificadas = int(marcas_unicas[mask_marca_real].nunique())
    else:
        mask_marca_real = pd.Series(False, index=df_res.index)
        n_filas_marca_real = 0
        n_marcas_identificadas = 0

    # ---- Bloque 3: Tarjetas de marca y características no numéricas ----
    # Características no numéricas = variables categóricas del maestro
    # (tipo de tecnología, fases, etc. — se excluyen las numéricas como
    # amperaje, voltaje, kVA).
    vars_cat = st.session_state.get("variables_categoricas", [])
    var_principal = st.session_state.get("var_principal_nombre", "")
    cols_caract = [c for c in vars_cat if c != var_principal and c in df_res.columns]
    if not cols_caract:
        cols_caract = [c for c in vars_cat if c in df_res.columns]

    if cols_caract:
        n_caract_por_fila = df_res[cols_caract].notna().sum(axis=1)
        n_con_marca_y_caract = int((mask_marca_real & (n_caract_por_fila >= 1)).sum())
        n_con_marca_y_2caract = int((mask_marca_real & (n_caract_por_fila >= 2)).sum())
        n_con_marca_y_3caract = int((mask_marca_real & (n_caract_por_fila >= 3)).sum())
    else:
        n_con_marca_y_caract = 0
        n_con_marca_y_2caract = 0
        n_con_marca_y_3caract = 0

    pct_marca_real = n_filas_marca_real / total
    pct_con_marca_y_caract = n_con_marca_y_caract / total
    pct_con_marca_y_2caract = n_con_marca_y_2caract / total
    pct_con_marca_y_3caract = n_con_marca_y_3caract / total

    st.markdown("#### 🏷️ Marcas y características identificadas")
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric(
        "🏷️ Marcas distintas",
        f"{n_marcas_identificadas:,}",
        help="Número de marcas únicas detectadas (excluye S/M, genéricas, etc.).",
    )
    c2.metric(
        "✅ % con marca real",
        f"{pct_marca_real:.1%}",
        help=f"{n_filas_marca_real:,} de {total:,} filas tienen una marca real (excluye genéricas, S/M y marca de componentes).",
    )
    c3.metric(
        "🔍 Marca + 1 característica",
        f"{pct_con_marca_y_caract:.1%}",
        help=f"{n_con_marca_y_caract:,} de {total:,} filas tienen marca real y al menos 1 característica no numérica.",
    )
    c4.metric(
        "🔍 Marca + 2 características",
        f"{pct_con_marca_y_2caract:.1%}",
        help=f"{n_con_marca_y_2caract:,} de {total:,} filas tienen marca real y al menos 2 características no numéricas.",
    )
    c5.metric(
        "🔍 Marca + 3 características",
        f"{pct_con_marca_y_3caract:.1%}",
        help=f"{n_con_marca_y_3caract:,} de {total:,} filas tienen marca real y al menos 3 características no numéricas.",
    )

    st.write("")

    # ---- Bloque 4: Top marcas detectadas ----
    if "Marca_Extraida" in df_res.columns:
        top_marcas = (
            marcas_unicas[mask_marca_real]
            .value_counts()
            .head(8)
            .sort_values(ascending=False)
        )
        if len(top_marcas) > 0:
            st.markdown("#### 🏆 Top marcas detectadas")
            st.caption("Las marcas reales más frecuentes en el archivo (excluye genéricas y sin marca).")
            import altair as alt

            df_top = top_marcas.reset_index()
            df_top.columns = ["Marca", "Cantidad"]
            chart = (
                alt.Chart(df_top)
                .mark_bar(color="#4C9AFF")
                .encode(
                    x=alt.X("Cantidad:Q", title="Filas"),
                    y=alt.Y(
                        "Marca:N",
                        sort="-x",
                        title=None,
                        axis=alt.Axis(labelLimit=200),
                    ),
                    tooltip=["Marca:N", "Cantidad:Q"],
                )
            )
            text = chart.mark_text(
                align="left",
                dx=4,
                color="#333333",
                fontSize=12,
            ).encode(text="Cantidad:Q")
            st.altair_chart((chart + text).properties(height=280), use_container_width=True)

    st.write("")

    d1, d2 = st.columns(2)
    with d1:
        if st.session_state.maestro_opt_data is not None:
            st.download_button(
                label=f"🧠 Descargar Maestro Optimizado",
                data=st.session_state.maestro_opt_data,
                file_name=f"Maestro_Optimizado_{st.session_state.linea_producto}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
                key="btn_descarga_maestro",
            )
        else:
            st.button("🧠 Maestro Optimizado (Sin aprendizajes nuevos)", disabled=True, width="stretch")

    with d2:
        if st.session_state.df_export_data is not None:
            st.download_button(
                label=f"� Descargar Resultado (Excel)",
                data=st.session_state.df_export_data,
                file_name=f"Resultado_{st.session_state.linea_producto}_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
                key="btn_descarga_resultado",
            )
        else:
            if st.button(
                "⚙️ Preparar Excel para descargar",
                width="stretch",
                key="btn_gen_resultado",
            ):
                with st.spinner("Generando Excel… Esto puede tardar unos segundos para archivos grandes."):
                    st.session_state.df_export_data = _generar_excel_resultado(
                        st.session_state.df_resultado,
                        st.session_state.kpis,
                        st.session_state.get("linea_producto", "Producto"),
                        st.session_state.get("archivo_origen", ""),
                        st.session_state.get("hoja_origen", ""),
                        st.session_state.get("modelo_ia_usado", ""),
                    )
                st.rerun()

# =====================================================================
# SECCIÓN 5: CREAR MAESTRO (DENTRO DEL TAB CREAR)
# =====================================================================
with tab_crear:
    from src.creador_maestro import (
        muestrear_veritrade,
        generar_maestro_con_ia,
        guardar_maestro_nuevo,
        MUESTRA_DEFAULT,
    )

    st.write("") # Espaciador
    st.markdown("### 🔧 Generador Automático de Maestros")
    st.caption(
        "Crea un maestro de clasificación a partir de un archivo Veritrade crudo. "
        "La IA analiza una muestra representativa y genera todas las hojas del maestro "
        "con marcas, características, patrones técnicos y reglas condicionales."
    )

    # ------------------------------------------------------------------
    # Session state para el generador
    # ------------------------------------------------------------------
    if "creador_step" not in st.session_state:
        st.session_state.creador_step = 0  # 0=upload, 1=form, 2=muestra, 3=generando, 4=resultado
    if "creador_muestra" not in st.session_state:
        st.session_state.creador_muestra = None
    if "creador_hojas" not in st.session_state:
        st.session_state.creador_hojas = None
    if "creador_maestro_bytes" not in st.session_state:
        st.session_state.creador_maestro_bytes = None
    if "creador_producto" not in st.session_state:
        st.session_state.creador_producto = ""
    if "creador_dominio" not in st.session_state:
        st.session_state.creador_dominio = {}
    if "creador_error" not in st.session_state:
        st.session_state.creador_error = None
    if "creador_progreso" not in st.session_state:
        st.session_state.creador_progreso = ""

    # ------------------------------------------------------------------
    # PASO 1: Subir Veritrade crudo + Template
    # ------------------------------------------------------------------
    st.markdown("#### 📁 Paso 1: Archivos de entrada")

    col_crudo, col_template = st.columns(2)

    with col_crudo:
        with st.container(border=True):
            st.markdown("**Veritrade crudo del PM**")
            st.caption("Archivo .xlsx con las importaciones de tu categoría.")
            archivo_crudo_creador = st.file_uploader(
                "Arrastra tu Veritrade .xlsx",
                type=["xlsx"],
                label_visibility="collapsed",
                key="up_crudo_creador",
            )
            if archivo_crudo_creador:
                try:
                    info_hojas_crudo = _listar_hojas_y_filas(archivo_crudo_creador.getvalue())
                    nombres_hojas_crudo = [i["nombre"] for i in info_hojas_crudo]
                    if nombres_hojas_crudo:
                        rec_crudo = _hoja_recomendada(info_hojas_crudo)
                        idx_crudo = nombres_hojas_crudo.index(rec_crudo) if rec_crudo in nombres_hojas_crudo else 0
                        hoja_crudo_creador = st.selectbox(
                            "Hoja de datos",
                            nombres_hojas_crudo,
                            index=idx_crudo,
                            key="hoja_crudo_creador",
                        )
                        filas_crudo = info_hojas_crudo[idx_crudo]["filas"]
                        st.caption(f"📄 **{hoja_crudo_creador}** — ~{filas_crudo:,} filas")
                    else:
                        hoja_crudo_creador = None
                        st.error("El archivo no tiene hojas.")
                except Exception as e:
                    hoja_crudo_creador = None
                    st.error(f"Error al leer: {e}")
            else:
                hoja_crudo_creador = None

    with col_template:
        with st.container(border=True):
            st.markdown("**Maestro plantilla (referencia)**")
            st.caption("Define el formato de salida. Sube tu plantilla .xlsx.")
            up_tpl = st.file_uploader(
                "Sube tu plantilla .xlsx",
                type=["xlsx"],
                label_visibility="collapsed",
                key="up_template",
            )
            template_bytes = None
            if up_tpl:
                up_tpl.seek(0)
                template_bytes = up_tpl.getvalue()
                st.success(f"✅ Plantilla: {up_tpl.name}")

    listo_paso1 = archivo_crudo_creador is not None and template_bytes is not None
    st.write("")

    # ------------------------------------------------------------------
    # PASO 2: Formulario de dominio del PM (COMPLETAMENTE OPCIONAL)
    # ------------------------------------------------------------------
    if listo_paso1:
        st.markdown("#### 🧠 Paso 2: Contexto de tu producto")
        st.caption(
            "Opcional: si tienes contexto del producto, la IA lo usará para mejorar las reglas. "
            "Si no填写 nada, la IA analizará directamente las descripciones."
        )

        c_producto, c_n_muestra = st.columns([2, 1])
        with c_producto:
            producto_nombre = st.text_input(
                "Nombre del producto a clasificar *",
                value=st.session_state.creador_producto or "",
                placeholder="Ej: Estabilizadores, Baterías, Cables...",
                help="Nombre corto de la categoría. Requerido para generar el maestro.",
            )
        with c_n_muestra:
            n_muestra = st.slider(
                "Tamaño de muestra",
                min_value=80,
                max_value=500,
                value=MUESTRA_DEFAULT,
                help="500 filas = máxima cobertura y calidad.",
            )

        with st.expander("📝 Contexto adicional del PM (todo opcional)", expanded=False):
            caracteristicas = st.text_area(
                "Características diferenciadoras (opcional)",
                value=st.session_state.creador_dominio.get("caracteristicas", ""),
                placeholder="Ej: Tecnología (online/trifásico), Fases, Formato (rack/piso), Gama...",
                height=70,
            )

            marcas_conocidas = st.text_area(
                "Marcas conocidas en esta categoría (opcional)",
                value=st.session_state.creador_dominio.get("marcas_conocidas", ""),
                placeholder="Ej: APC, Eaton, CyberPower, Lestar, Schneider...",
                height=60,
            )

            patrones_tecnicos = st.text_area(
                "Patrones numéricos técnicos relevantes (opcional)",
                value=st.session_state.creador_dominio.get("patrones_tecnicos", ""),
                placeholder="Ej: Voltaje (110V, 220V), Potencia (kVA, KW), Capacidad (Ah, Wh)...",
                height=60,
            )

        # Dos botones: uno con contexto, otro directo (sin formulario)
        c_directo, c_contexto = st.columns([1, 1])
        with c_directo:
            if st.button(
                "⚡ Generar directamente (sin contexto extra)",
                type="secondary",
                width="stretch",
                disabled=not producto_nombre.strip(),
            ):
                if not producto_nombre.strip():
                    st.error("⚠️ Escribe el nombre del producto.")
                else:
                    st.session_state.creador_producto = producto_nombre.strip()
                    st.session_state.creador_dominio = {
                        "caracteristicas": "",
                        "marcas_conocidas": "",
                        "patrones_tecnicos": "",
                    }
                    st.session_state.creador_step = 1
        with c_contexto:
            if st.button(
                "✅ Usar contexto y ver muestra",
                type="primary",
                width="stretch",
                disabled=not producto_nombre.strip(),
            ):
                if not producto_nombre.strip():
                    st.error("⚠️ Escribe el nombre del producto.")
                else:
                    st.session_state.creador_producto = producto_nombre.strip()
                    st.session_state.creador_dominio = {
                        "caracteristicas": caracteristicas.strip(),
                        "marcas_conocidas": marcas_conocidas.strip(),
                        "patrones_tecnicos": patrones_tecnicos.strip(),
                    }
                    st.session_state.creador_step = 1

    # ------------------------------------------------------------------
    # PASO 3: Muestra estratificada
    # ------------------------------------------------------------------
    if listo_paso1 and st.session_state.creador_step >= 1 and st.session_state.creador_producto:
        st.markdown("#### 📊 Paso 3: Muestra representativa")
        st.caption("Muestreo estratificado determinístico — sin llamar a la IA, solo pandas.")

        try:
            archivo_crudo_creador.seek(0)
            df_crudo_completo = pd.read_excel(
                BytesIO(archivo_crudo_creador.getvalue()),
                sheet_name=hoja_crudo_creador,
            )
            df_muestra, df_dedup = muestrear_veritrade(
                df_crudo_completo,
                n_muestra=n_muestra,
            )
            st.session_state.creador_muestra = df_muestra
            st.session_state.creador_step = 2

            c_info1, c_info2, c_info3 = st.columns(3)
            c_info1.metric("Filas en archivo", f"{len(df_crudo_completo):,}")
            c_info2.metric("Únicas (dedup)", f"{len(df_dedup):,}")
            c_info3.metric("Muestra seleccionada", f"{len(df_muestra):,}")

            with st.expander(f"👁️ Ver muestra ({len(df_muestra)} filas)", expanded=False):
                cols_desc_muestra = identificar_columnas_descripcion(df_muestra.columns)
                cols_mostrar = cols_desc_muestra[:6] if cols_desc_muestra else list(df_muestra.columns[:6])
                st.dataframe(df_muestra[cols_mostrar].head(50), width="stretch", hide_index=True)
                if len(df_muestra) > 50:
                    st.caption(f"Mostrando 50 de {len(df_muestra)} filas de la muestra.")

        except Exception as e:
            st.error(f"Error al generar la muestra: {e}")
            st.session_state.creador_error = str(e)

    # ------------------------------------------------------------------
    # PASO 4: Generar maestro con IA
    # ------------------------------------------------------------------
    if (
        st.session_state.creador_step >= 2
        and st.session_state.creador_muestra is not None
        and template_bytes is not None
    ):
        st.markdown("#### 🤖 Paso 4: Generar maestro con IA")
        st.caption("Una sola llamada a Gemini genera todas las hojas del maestro.")

        # Configuración de IA (reutilizar key del tab de clasificación si existe)
        api_key_creador = _obtener_api_key_de_secrets()
        usar_ia_creador = GENAI_DISPONIBLE

        if not usar_ia_creador:
            st.error("⚠️ google-genai no está instalado.")
        else:
            c_ak, c_modelo_gen = st.columns([2, 1])
            with c_ak:
                api_key_creador = st.text_input(
                    "API Key de Gemini",
                    type="password",
                    value=api_key_creador,
                    key="api_key_creador",
                    help="Requerida para generar el maestro. Se toma de secrets.toml si existe.",
                ).strip()
            with c_modelo_gen:
                modelo_gen = st.selectbox(
                    "Modelo",
                    config.MODELOS_IA_DISPONIBLES,
                    index=0,
                    key="modelo_gen",
                )

            listo_generar = bool(api_key_creador)

            if st.button(
                "🚀 Generar Maestro con IA",
                type="primary",
                width="stretch",
                disabled=not listo_generar,
            ):
                if not api_key_creador:
                    st.warning("Ingresa una API Key.")
                else:
                    st.session_state.creador_step = 3
                    st.session_state.creador_error = None

                    # Llamada a la IA (síncrona con spinner)
                    with st.spinner("🔍 Leyendo template..."):
                        pass  # Ya leído

                    def _cb_creador(fase, msg):
                        st.session_state.creador_progreso = f"[{fase}] {msg}"

                    try:
                        with st.spinner("🤖 Generando maestro con IA... Esto puede tomar 30-60 segundos."):
                            hojas_generadas = generar_maestro_con_ia(
                                api_key=api_key_creador,
                                modelo=modelo_gen,
                                df_muestra=st.session_state.creador_muestra,
                                ruta_template=BytesIO(template_bytes),
                                producto=st.session_state.creador_producto,
                                dominio=st.session_state.creador_dominio,
                                progreso_callback=_cb_creador,
                            )

                        st.session_state.creador_hojas = hojas_generadas
                        st.session_state.creador_step = 4
                        st.success("✅ Maestro generado exitosamente. Revisa el resultado abajo.")

                    except Exception as e:
                        st.session_state.creador_error = str(e)
                        st.session_state.creador_step = 2
                        st.error(f"❌ Error al generar: {e}")

    # ------------------------------------------------------------------
    # PASO 5: Revisión y descarga
    # ------------------------------------------------------------------
    if (
        st.session_state.creador_step >= 4
        and st.session_state.creador_hojas is not None
    ):
        st.markdown("#### 📥 Paso 5: Revisión y descarga")
        st.success("✅ Maestro generado. Revisa cada hoja antes de usarlo en producción.")

        hojas = st.session_state.creador_hojas

        # KPIs del maestro
        n_marcas = len(hojas.get("1_Marcas", []))
        n_stopwords = len(hojas.get("1b_Palabras_Ignorar", []))
        n_carac = len(hojas.get("2_Caracteristicas", []))
        n_potencia = len(hojas.get("3_Tecnico_Potencia_NOEDIT", []))
        n_regex = len(hojas.get("4_Tecnico_RegexMarca_NOEDIT", []))
        n_cond = len(hojas.get("5_Condicionales", []))

        km1, km2, km3, km4, km5, km6 = st.columns(6)
        km1.metric("🏷️ Marcas", n_marcas)
        km2.metric("🚫 Stopwords", n_stopwords)
        km3.metric("📋 Características", n_carac)
        km4.metric("⚡ Potencia", n_potencia)
        km5.metric("🔍 Regex", n_regex)
        km6.metric("🔀 Condicionales", n_cond)

        st.write("")

        # Vista previa por hoja
        tab_names = [f"1_Marcas ({n_marcas})",
                     f"1b_Stopwords ({n_stopwords})",
                     f"2_Características ({n_carac})",
                     f"3_Potencia ({n_potencia})",
                     f"4_Regex ({n_regex})",
                     f"5_Condicionales ({n_cond})"]
        tabs_hojas = st.tabs(tab_names)

        hojas_keys = [
            "1_Marcas", "1b_Palabras_Ignorar", "2_Caracteristicas",
            "3_Tecnico_Potencia_NOEDIT", "4_Tecnico_RegexMarca_NOEDIT", "5_Condicionales",
        ]

        for tab_h, key_h in zip(tabs_hojas, hojas_keys):
            with tab_h:
                df_h = hojas.get(key_h)
                if df_h is not None and len(df_h) > 0:
                    st.dataframe(df_h, width="stretch", hide_index=True)
                else:
                    st.info("Esta hoja está vacía (la IA no encontró datos relevantes para esta sección).")

        st.write("")

        # Descarga
        try:
            bytes_maestro_nuevo = guardar_maestro_nuevo(
                hojas=hojas,
                producto=st.session_state.creador_producto,
            )
            st.session_state.creador_maestro_bytes = bytes_maestro_nuevo

            sufijo = datetime.now().strftime("%Y-%m-%d")
            nombre_archivo = f"Maestro_{st.session_state.creador_producto}_v1_{sufijo}.xlsx"

            st.download_button(
                label=f"📥 Descargar {nombre_archivo}",
                data=bytes_maestro_nuevo,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                width="stretch",
                key="btn_descarga_maestro_creador",
            )

            st.caption(
                "💡 **Siguiente paso:** Sube este maestro en el tab 📊 Clasificar "
                "para probarlo contra tu Veritrade completo."
            )

        except Exception as e:
            st.error(f"Error al generar el Excel: {e}")

    # Botón para reiniciar
    if st.session_state.creador_step >= 4:
        st.write("")
        if st.button("🔄 Crear otro maestro", width="stretch"):
            st.session_state.creador_step = 0
            st.session_state.creador_muestra = None
            st.session_state.creador_hojas = None
            st.session_state.creador_maestro_bytes = None
            st.session_state.creador_error = None
            st.rerun()