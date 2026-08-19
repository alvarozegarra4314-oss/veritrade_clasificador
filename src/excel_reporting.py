# -*- coding: utf-8 -*-
"""Generacion de reportes Excel con resultados y resumen ejecutivo."""

from __future__ import annotations

import os
import re
import tempfile
from io import BytesIO
from pathlib import Path
from typing import IO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

COLOR_HEADER = "1F4E78"
COLOR_SUBHEADER = "D9EAF7"
COLOR_TEXT = "FFFFFF"
COLOR_BODY = "FFFFFF"
COLOR_BORDER = "D9D9D9"
BORDER = Border(*(Side(style="thin", color=COLOR_BORDER) for _ in range(4)))


def _nombre_columna(columnas, patrones: tuple[str, ...]) -> str | None:
    for columna in columnas:
        normalizada = re.sub(r"[^A-Z0-9]", "", str(columna).upper())
        if any(patron in normalizada for patron in patrones):
            return columna
    return None


def _numero(serie: pd.Series) -> pd.Series:
    """Convierte valores numericos con formatos comunes de Excel/Veritrade."""
    if pd.api.types.is_numeric_dtype(serie):
        return pd.to_numeric(serie, errors="coerce").fillna(0)
    texto = serie.astype(str).str.strip()
    tiene_coma = texto.str.contains(",", regex=False)
    texto = texto.where(
        ~tiene_coma,
        texto.str.replace(".", "", regex=False).str.replace(",", ".", regex=False),
    )
    return pd.to_numeric(texto, errors="coerce").fillna(0)


def _normalizar_columnas(df: pd.DataFrame) -> pd.DataFrame:
    """Garantiza encabezados no vacios y unicos, requisito de Excel Table."""
    resultado = df.copy()
    nombres = []
    vistos: dict[str, int] = {}
    for indice, columna in enumerate(resultado.columns, start=1):
        base = str(columna).strip() or f"Columna_{indice}"
        vistos[base] = vistos.get(base, 0) + 1
        nombres.append(base if vistos[base] == 1 else f"{base}_{vistos[base]}")
    resultado.columns = nombres
    return resultado


def _crear_tabla(ws, nombre: str, inicio: int, fin: int, columnas: int) -> None:
    if fin < inicio or columnas < 1:
        return
    ref = f"A{inicio}:{get_column_letter(columnas)}{fin}"
    tabla = Table(displayName=nombre[:255], ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(tabla)


def _formatear_rango(ws, fila_inicio: int, fila_fin: int, columnas: int, encabezado: bool = False) -> None:
    for fila in range(fila_inicio, fila_fin + 1):
        for columna in range(1, columnas + 1):
            celda = ws.cell(fila, columna)
            celda.border = BORDER
            celda.alignment = Alignment(vertical="center", wrap_text=False)
            if encabezado and fila == fila_inicio:
                celda.fill = PatternFill("solid", fgColor=COLOR_HEADER)
                celda.font = Font(bold=True, color=COLOR_TEXT)


def _es_caracteristica(columna: str, marca_columna: str | None) -> bool:
    nombre = str(columna).upper()
    excluidas = {"DESCRIPCION", "PRODUCTO_TEXTO", "MODELO_SERIE", "ORIGEN", "RESCATADO", "ES_PRODUCTO"}
    if columna == marca_columna or any(palabra in nombre for palabra in excluidas):
        return False
    return any(palabra in nombre for palabra in ("PRODUCTO", "TIPO", "SALIDA", "TECNOLOGIA", "FASES", "POTENCIA", "KVA", "TENSION", "VOLTAJE"))


def _resumen_marcas(df: pd.DataFrame, marca_columna: str | None, qty_columna: str | None, fob_columna: str | None) -> pd.DataFrame:
    if not marca_columna:
        return pd.DataFrame(columns=["Marca", "Registros", "Qty2_Total", "FOB_Total", "Participacion"])
    trabajo = pd.DataFrame({"Marca": df[marca_columna].fillna("Sin marca").replace("", "Sin marca")})
    trabajo["Registros"] = 1
    if qty_columna:
        trabajo["Qty2_Total"] = _numero(df[qty_columna])
    if fob_columna:
        trabajo["FOB_Total"] = _numero(df[fob_columna])
    resumen = trabajo.groupby("Marca", as_index=False).sum(numeric_only=True).sort_values("Registros", ascending=False)
    resumen["Participacion"] = resumen["Registros"] / max(len(df), 1)
    resumen.insert(0, "Ranking", range(1, len(resumen) + 1))
    return resumen


def _resumen_caracteristicas(df: pd.DataFrame, marca_columna: str | None, qty_columna: str | None, fob_columna: str | None) -> pd.DataFrame:
    columnas = [columna for columna in df.columns if _es_caracteristica(columna, marca_columna)]
    filas = []
    for columna in columnas:
        trabajo = pd.DataFrame({"Caracteristica": columna, "Valor": df[columna].fillna("Sin valor").replace("", "Sin valor")})
        trabajo["Registros"] = 1
        if qty_columna:
            trabajo["Qty2_Total"] = _numero(df[qty_columna])
        if fob_columna:
            trabajo["FOB_Total"] = _numero(df[fob_columna])
        filas.append(trabajo.groupby(["Caracteristica", "Valor"], as_index=False).sum(numeric_only=True))
    if not filas:
        return pd.DataFrame(columns=["Caracteristica", "Valor", "Registros", "Qty2_Total", "FOB_Total"])
    return pd.concat(filas, ignore_index=True).sort_values("Registros", ascending=False)


def _escribir_seccion(ws, fila: int, titulo: str, subtitulo: str, datos: pd.DataFrame, nombre_tabla: str) -> int:
    ws.cell(fila, 1, titulo).font = Font(bold=True, size=13, color=COLOR_HEADER)
    ws.cell(fila + 1, 1, subtitulo).font = Font(italic=True, color="666666")
    inicio = fila + 3
    for columna, nombre in enumerate(datos.columns, start=1):
        ws.cell(inicio, columna, nombre)
    for indice, valores in enumerate(datos.itertuples(index=False, name=None), start=inicio + 1):
        for columna, valor in enumerate(valores, start=1):
            ws.cell(indice, columna, valor)
    fin = inicio + len(datos)
    _formatear_rango(ws, inicio, fin, len(datos.columns), encabezado=True)
    if len(datos):
        _crear_tabla(ws, nombre_tabla, inicio, fin, len(datos.columns))
    return fin + 3


def _generar_libro(df: pd.DataFrame, destino) -> None:
    df = _normalizar_columnas(df)
    marca_columna = _nombre_columna(df.columns, ("MARCAEXTRAIDA", "MARCADECLARADA", "MARCAESTANDARIZADA", "MARCA"))
    qty_columna = _nombre_columna(df.columns, ("QTY2", "CANTIDAD", "QUANTITY"))
    fob_columna = _nombre_columna(df.columns, ("FOBTOTAL", "FOB"))
    resumen_marcas = _resumen_marcas(df, marca_columna, qty_columna, fob_columna)
    resumen_caracteristicas = _resumen_caracteristicas(df, marca_columna, qty_columna, fob_columna)

    with pd.ExcelWriter(destino, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Clasificacion")

    libro = load_workbook(destino)
    resultados = libro["Clasificacion"]
    resultados.sheet_view.showGridLines = False
    resultados.freeze_panes = "A2"
    _formatear_rango(resultados, 1, max(1, resultados.max_row), resultados.max_column, encabezado=True)
    _crear_tabla(resultados, "Tabla_Clasificacion", 1, resultados.max_row, resultados.max_column)

    resumen = libro.create_sheet("Resumen Ejecutivo", 0)
    resumen.sheet_view.showGridLines = False
    resumen.freeze_panes = "A4"
    resumen["A1"] = "Resumen Ejecutivo de Clasificacion"
    resumen["A1"].font = Font(bold=True, size=18, color=COLOR_HEADER)
    resumen["A2"] = f"Registros analizados: {len(df):,} | Marca: {marca_columna or 'no disponible'} | Qty2: {qty_columna or 'no disponible'} | FOB: {fob_columna or 'no disponible'}"
    resumen["A2"].font = Font(italic=True, color="666666")

    fila = 4
    fila = _escribir_seccion(resumen, fila, "Ranking y participacion por marca", "Cantidad de registros, volumen, FOB total y participacion sobre el total.", resumen_marcas, "Tabla_Resumen_Marcas")
    _escribir_seccion(resumen, fila, "Resumen por caracteristica", "Distribucion de valores clasificados con conteo, Qty2 y FOB cuando existen en el resultado.", resumen_caracteristicas, "Tabla_Resumen_Caracteristicas")

    for hoja in libro.worksheets:
        for columna in range(1, hoja.max_column + 1):
            valores = [len(str(hoja.cell(fila, columna).value or "")) for fila in range(1, min(hoja.max_row, 1000) + 1)]
            hoja.column_dimensions[get_column_letter(columna)].width = min(max(max(valores, default=10) + 3, 12), 55)
        hoja.sheet_view.showGridLines = False
    libro.save(destino)


def generar_reporte_excel(df: pd.DataFrame, destino) -> None:
    """Genera el Excel final en una ruta o BytesIO, sin alterar el DataFrame."""
    if df is None or df.empty:
        raise ValueError("No hay datos para generar el reporte Excel.")
    if isinstance(destino, BytesIO):
        _generar_libro(df, destino)
        destino.seek(0)
        return
    ruta = Path(destino)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    fd, temporal = tempfile.mkstemp(suffix=".xlsx", dir=ruta.parent)
    os.close(fd)
    try:
        _generar_libro(df, temporal)
        os.replace(temporal, ruta)
    finally:
        if os.path.exists(temporal):
            os.remove(temporal)
