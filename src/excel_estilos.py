# -*- coding: utf-8 -*-
"""
Módulo de formateo y estilos visuales para archivos de Excel (.xlsx)
---------------------------------------------------------------------
Centraliza la presentación profesional de reportes y maestros:
  - Formato de tabla nativa (Ctrl+T)
  - Desactivación de líneas de cuadrícula (gridlines) para acabado prolijo
  - Fondo blanco y encabezados con color institucional
  - Exclusión de hojas de instrucciones o documentales de formato tabla
  - Reglas de color por estado (Auditoría/Log)
"""

from __future__ import annotations

import re
import copy
import os
import tempfile
from typing import Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Colores predeterminados (Institucionales / Profesionales)
COLOR_HEADER_BG = "1F4E78"       # Azul oscuro corporativo
COLOR_HEADER_TEXT = "FFFFFF"     # Texto blanco
COLOR_BODY_BG = "FFFFFF"         # Fondo blanco prolijo
COLOR_BORDER = "D9D9D9"          # Borde gris sutil

# Colores para estados del Log de Auditoría
COLOR_LOG_OK = "E2EFDA"          # Verde claro (Agregado)
COLOR_LOG_OMITIDO = "FFF2CC"     # Amarillo claro (Omitido)
COLOR_LOG_PENDIENTE = "FCE4E4"   # Rojo claro (Pendiente revisión)

_BORDE_FINO = Side(style="thin", color=COLOR_BORDER)
_BORDE_CELDA = Border(left=_BORDE_FINO, right=_BORDE_FINO, top=_BORDE_FINO, bottom=_BORDE_FINO)


def _color_fila_log(estado: str) -> Optional[str]:
    if not estado:
        return None
    estado_up = str(estado).upper()
    if estado_up.startswith("AGREGADO"):
        return COLOR_LOG_OK
    if estado_up.startswith("OMITIDO"):
        return COLOR_LOG_OMITIDO
    if estado_up.startswith("PENDIENTE"):
        return COLOR_LOG_PENDIENTE
    return None


def es_hoja_instrucciones(nombre_hoja: str) -> bool:
    """Identifica si una hoja es de instrucciones/documentación para no convertirla en tabla nativa."""
    nombre_up = nombre_hoja.upper()
    keywords = ["INSTRUCCION", "INSTRUCCIONES", "HELP", "DOC", "DOCUMENTACION", "READ_ME", "README"]
    return any(kw in nombre_up for kw in keywords)


def aplicar_estilo_hoja_excel(ws, df: pd.DataFrame, es_log: bool = False) -> None:
    """
    Aplica formato visual profesional a una hoja de openpyxl:
      1. Desactiva gridlines (showGridLines = False) para vista limpia.
      2. Si NO es hoja de instrucciones, crea una tabla nativa de Excel (Ctrl+T).
      3. Aplica encabezado con color predeterminado (azul oscuro + texto blanco).
      4. Aplica fondo blanco a las celdas de datos (o color por estado en el log).
      5. Ajusta ancho de columnas automáticamente.
      6. Congela el panel en la primera fila (encabezado siempre visible).
    """
    if ws.max_row == 0 or ws.max_column == 0:
        return

    n_filas = len(df)
    n_cols = len(df.columns)

    # En hojas grandes el relleno/borde celda a celda es O(filas*cols) y puede
    # tardar decenas de segundos (ej. resultado de clasificación con miles de
    # filas). El fondo blanco ya está garantizado al desactivar gridlines y por
    # el estilo de tabla nativa, así que en hojas grandes se omite ese paso sin
    # perder el acabado visual (encabezado, tabla, filtros y paneles congelados
    # se aplican igual).
    UMBRAL_RELLENO_CELDA_A_CELDA = 2000
    aplicar_cuerpo = n_filas <= UMBRAL_RELLENO_CELDA_A_CELDA

    # 1. Desactivar líneas de cuadrícula para un aspecto limpio y profesional
    ws.sheet_view.showGridLines = False

    # 2. Encabezados
    for col_idx in range(1, n_cols + 1):
        celda = ws.cell(row=1, column=col_idx)
        celda.font = Font(bold=True, color=COLOR_HEADER_TEXT)
        celda.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        celda.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        celda.border = _BORDE_CELDA

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24

    # 3. Filas de datos (solo en hojas chicas; en grandes se omite por rendimiento)
    col_estado_idx = None
    if es_log and "Estado" in df.columns:
        col_estado_idx = list(df.columns).index("Estado") + 1

    if aplicar_cuerpo:
        for fila_idx in range(2, n_filas + 2):
            color_fila = None
            if col_estado_idx:
                estado_val = ws.cell(row=fila_idx, column=col_estado_idx).value
                color_fila = _color_fila_log(estado_val)

            # Si no tiene un color especial de log, el fondo es blanco limpio
            if color_fila is None:
                color_fila = COLOR_BODY_BG

            for col_idx in range(1, n_cols + 1):
                celda = ws.cell(row=fila_idx, column=col_idx)
                celda.border = _BORDE_CELDA
                celda.alignment = Alignment(vertical="center", wrap_text=False)
                celda.fill = PatternFill("solid", fgColor=color_fila)

    # 4. Autofiltro solo para hojas de datos. Las instrucciones no son una
    # tabla y no deben llevar filtros ni referencias de tabla.
    es_instrucciones = es_hoja_instrucciones(ws.title)
    if not es_instrucciones:
        ws.auto_filter.ref = ws.dimensions
    else:
        ws.auto_filter.ref = None

    # 5. Crear Tabla Nativa de Excel (Ctrl+T) solo si NO es hoja de instrucciones
    if not es_instrucciones and ws.max_row >= 2 and ws.max_column >= 1 and not ws.tables:
        ref = f"A1:{get_column_letter(n_cols)}{n_filas + 1}"
        nombre_limpio = re.sub(r"[^A-Za-z0-9_]", "_", ws.title)
        nombre_tabla = f"Tabla_{nombre_limpio}"[:240]

        tabla = Table(displayName=nombre_tabla, ref=ref)
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1",  # Estilo limpio sobre fondo blanco
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,     # Mantener fondo blanco prolijo
            showColumnStripes=False,
        )
        ws.add_table(tabla)

    # 6. Ancho de columnas ajustado dinámicamente al contenido
    for col_idx, col_name in enumerate(df.columns, start=1):
        letra = get_column_letter(col_idx)
        try:
            max_contenido = df[col_name].astype(str).map(len).max()
        except Exception:
            max_contenido = 10
        max_contenido = max_contenido if pd.notna(max_contenido) else 10
        ancho = min(max(len(str(col_name)), int(max_contenido)) + 3, 60)
        ws.column_dimensions[letra].width = ancho


def restaurar_hoja_instrucciones(ruta_salida, ruta_original) -> None:
    """Restaura la hoja de instrucciones original sin pasarla por pandas.

    La hoja contiene texto, espacios y estilos que no deben interpretarse
    como encabezados de datos. Se guarda mediante un archivo temporal para
    evitar dejar un XLSX incompleto si el proceso se interrumpe.
    """
    libro_salida = load_workbook(ruta_salida)
    libro_original = load_workbook(ruta_original)
    nombre = next((n for n in libro_original.sheetnames if es_hoja_instrucciones(n)), None)
    if not nombre or nombre not in libro_salida.sheetnames:
        return

    hoja_original = libro_original[nombre]
    hoja_salida = libro_salida[nombre]
    indice = libro_salida.index(hoja_salida)
    libro_salida.remove(hoja_salida)
    hoja_nueva = libro_salida.create_sheet(nombre, indice)

    for fila in hoja_original.iter_rows():
        for celda_original in fila:
            celda_nueva = hoja_nueva[celda_original.coordinate]
            celda_nueva.value = celda_original.value
            if celda_original.has_style:
                # No copiar _style entre libros: sus índices internos no
                # pertenecen al libro destino y Excel intenta repararlos.
                celda_nueva.font = copy.copy(celda_original.font)
                celda_nueva.fill = copy.copy(celda_original.fill)
                celda_nueva.border = copy.copy(celda_original.border)
                celda_nueva.alignment = copy.copy(celda_original.alignment)
                celda_nueva.protection = copy.copy(celda_original.protection)
            if celda_original.number_format:
                celda_nueva.number_format = celda_original.number_format
            if celda_original.hyperlink:
                celda_nueva._hyperlink = copy.copy(celda_original.hyperlink)
            if celda_original.comment:
                celda_nueva.comment = copy.copy(celda_original.comment)

    for rango in hoja_original.merged_cells.ranges:
        hoja_nueva.merge_cells(str(rango))
    for letra, dimension in hoja_original.column_dimensions.items():
        hoja_nueva.column_dimensions[letra] = copy.copy(dimension)
    for numero, dimension in hoja_original.row_dimensions.items():
        hoja_nueva.row_dimensions[numero] = copy.copy(dimension)

    hoja_nueva.sheet_view.showGridLines = False
    hoja_nueva.auto_filter.ref = None
    hoja_nueva.freeze_panes = None

    carpeta = os.path.dirname(os.path.abspath(str(ruta_salida)))
    fd, temporal = tempfile.mkstemp(suffix=".xlsx", dir=carpeta)
    os.close(fd)
    try:
        libro_salida.save(temporal)
        os.replace(temporal, ruta_salida)
    finally:
        if os.path.exists(temporal):
            os.remove(temporal)
